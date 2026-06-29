"""
Demo de la interfaz — Simula el hardware de la Raspberry Pi
para poder visualizar la UI en Windows.
NO MODIFICA EL CÓDIGO ORIGINAL. Solo para previsualización.
"""
import tkinter as tk
from tkinter import messagebox
import threading
import time
import sqlite3
import os
import math
import random
from datetime import datetime

# --- IMPORTACIONES PARA LA GRÁFICA ---
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure

# --- IMPORTACIONES PARA LOGOS ---
try:
    from PIL import Image, ImageTk
    PIL_DISPONIBLE = True
    try:
        RESAMPLE_METHOD = Image.LANCZOS
    except AttributeError:
        RESAMPLE_METHOD = Image.ANTIALIAS
except ImportError:
    PIL_DISPONIBLE = False

# --- PALETA DE COLORES WKK ---
COLOR_VERDE_WKK = "#0E8A3E"
COLOR_VERDE_OSCURO = "#0A6B2F"
COLOR_VERDE_CLARO = "#E8F5E9"
COLOR_FONDO = "#F0F2F5"
COLOR_TARJETA = "#FFFFFF"
COLOR_BORDE = "#DEE2E6"
COLOR_TEXTO = "#1A1A2E"
COLOR_TEXTO_SEC = "#6C757D"
COLOR_OK = "#00C853"
COLOR_NOK = "#E53935"

# --- SIMULACIÓN DE HARDWARE (para demo en Windows) ---
class LEDSimulado:
    def __init__(self, pin): self.pin = pin; self.estado = False
    def on(self): self.estado = True
    def off(self): self.estado = False
    def close(self): pass

class SensorSimulado:
    def __init__(self, pin, pull_up=False): self.pin = pin; self.when_pressed = None
    def close(self): pass

led = LEDSimulado(17)
sensor_pulso = SensorSimulado(4, pull_up=False)

# --- VARIABLES GLOBALES DE CONTROL ---
hilo_activo = True
fuerza_actual = 0.0
esperando_corte = False
timer_activo = False
tiempo_timer = 2.0
limite_ok = 598.2  # 61 kg en Newtons (61 * 9.80665)
KG_A_N = 9.80665
UMBRAL_ACTIVACION_N = 10.0 * KG_A_N  # 10 kg en Newtons
LIMITE_PICO_N = 90.0 * KG_A_N       # 90 kg en Newtons
muestras_timer = []

evento_corte_ui = False
tipo_corte = ""
valor_corte = 0.0
piezas_ok = 0
piezas_nok = 0

# --- VARIABLES PARA EL HISTORIAL DE LA GRÁFICA ---
MAX_PUNTOS = 100  
datos_fuerza = [0.0] * MAX_PUNTOS
datos_limite = [598.2] * MAX_PUNTOS

# --- SIMULACIÓN DE MODBUS ---
modbus_conectado = True  # Simulado

# --- CALIBRACIÓN MATEMÁTICA ---
LECTURA_CERO = 7.43     
LECTURA_MAXIMA = 8.23   
FUERZA_MAXIMA = 100.0   

# --- RUTAS ---
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ASSETS_DIR = os.path.join(SCRIPT_DIR, "assets")
DB_PATH = os.path.join(SCRIPT_DIR, "registros_prensa_demo.db")

# --- BASE DE DATOS SQLite ---
def inicializar_db():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS registros (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha TEXT NOT NULL,
            hora TEXT NOT NULL,
            valor_fuerza REAL NOT NULL,
            resultado TEXT NOT NULL,
            umbral REAL,
            minimo_ok REAL,
            maximo_ok REAL
        )
    """)
    conn.commit()
    conn.close()

inicializar_db()

def registrar_corte(valor, resultado, umbral, minimo, maximo):
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        ahora = datetime.now()
        cursor.execute(
            "INSERT INTO registros (fecha, hora, valor_fuerza, resultado, umbral, minimo_ok, maximo_ok) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (ahora.strftime("%Y-%m-%d"), ahora.strftime("%H:%M:%S"), valor, resultado, umbral, minimo, maximo)
        )
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"Error al registrar en DB: {e}")

def exportar_a_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM registros ORDER BY id")
        registros = cursor.fetchall()
        conn.close()
        
        if not registros:
            messagebox.showinfo("Exportar Excel", "No hay registros para exportar.")
            return
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Registros de Prensa"
        
        header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='0E8A3E', end_color='0E8A3E', fill_type='solid')
        ok_fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
        nok_fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
        center = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        headers = ['ID', 'Fecha', 'Hora', 'Valor Fuerza', 'Resultado', 'Umbral', 'Mínimo OK', 'Máximo OK']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border
        
        for row_idx, registro in enumerate(registros, 2):
            for col_idx, val in enumerate(registro, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = center
                cell.border = thin_border
                if col_idx == 5:
                    cell.fill = ok_fill if val == "OK" else nok_fill
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 15
        
        fecha_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = os.path.join(SCRIPT_DIR, f"registros_prensa_{fecha_str}.xlsx")
        wb.save(filename)
        messagebox.showinfo("Exportar Excel", f"Archivo exportado exitosamente:\n{filename}")
        
    except ImportError:
        messagebox.showerror("Error", "Se requiere instalar openpyxl:\npip install openpyxl")
    except Exception as e:
        messagebox.showerror("Error", f"Error al exportar:\n{e}")


# --- 4. HILO DE LECTURA SIMULADO ---
_sim_t = 0.0
def tarea_modbus_alta_velocidad():
    global fuerza_actual, esperando_corte, evento_corte_ui
    global tipo_corte, valor_corte, piezas_ok, piezas_nok, _sim_t
    global timer_activo, muestras_timer
    
    start_timer_time = 0.0
    
    while hilo_activo:
        _sim_t += 0.05
        # Simular lectura de fuerza en kg y convertir a N
        if not esperando_corte:
            # Fuerza en reposo es casi cero (ruido menor a 1 kg)
            f_calc_kg = random.uniform(0.0, 1.0)
        elif esperando_corte and not timer_activo:
            # Pistón bajando, la fuerza sube rápido
            f_calc_kg = 5.0 + (_sim_t % 5.0) * 15.0
            if f_calc_kg > 65.0:
                f_calc_kg = 65.0
        else:
            # Timer activo: fuerza estable cerca del límite (63 kg)
            f_calc_kg = 63.0 + random.uniform(-0.5, 0.5)
            # Meter un pico transitorio ocasional > 90 kg
            if random.random() < 0.1:
                f_calc_kg = 95.0
                
        if f_calc_kg < 0: f_calc_kg = 0.0
        if f_calc_kg > FUERZA_MAXIMA: f_calc_kg = FUERZA_MAXIMA
        
        fuerza_actual = f_calc_kg * KG_A_N
        
        # --- EVALUACIÓN A VELOCIDAD DE HARDWARE ---
        if esperando_corte:
            if not timer_activo:
                if fuerza_actual >= UMBRAL_ACTIVACION_N:
                    timer_activo = True
                    start_timer_time = time.time()
                    muestras_timer = [fuerza_actual]
            else:
                muestras_timer.append(fuerza_actual)
                if time.time() - start_timer_time >= tiempo_timer:
                    led.off()
                    
                    # Tomar las últimas 8 a 10 lecturas
                    lecturas_estables = muestras_timer[-10:] if len(muestras_timer) >= 10 else muestras_timer
                    
                    # Filtrar picos > 90 kg (LIMITE_PICO_N)
                    picos = [val for val in lecturas_estables if val > LIMITE_PICO_N]
                    if len(lecturas_estables) > 0 and (len(picos) / len(lecturas_estables)) < 0.5:
                        lecturas_filtradas = [val for val in lecturas_estables if val <= LIMITE_PICO_N]
                    else:
                        lecturas_filtradas = lecturas_estables
                    
                    if lecturas_filtradas:
                        valor_corte = sum(lecturas_filtradas) / len(lecturas_filtradas)
                    else:
                        valor_corte = fuerza_actual
                        
                    if valor_corte >= limite_ok:
                        piezas_ok += 1
                        tipo_corte = "OK"
                    else:
                        piezas_nok += 1
                        tipo_corte = "NOK"
                    
                    registrar_corte(valor_corte, tipo_corte, UMBRAL_ACTIVACION_N, limite_ok, None)
                        
                    esperando_corte = False
                    timer_activo = False
                    evento_corte_ui = True
                
        time.sleep(0.05)


# --- 5. FUNCIONES DE INTERFAZ Y PULSO ---
def encender_led_manual():
    led.on()
    canvas.itemconfig(indicador_led, fill=COLOR_OK)
    estado_label.config(text="LED ENCENDIDO MANUAL")

def apagar_led_manual():
    global esperando_corte, timer_activo
    led.off()
    canvas.itemconfig(indicador_led, fill=COLOR_NOK)
    canvas.itemconfig(indicador_pulso, fill=COLOR_BORDE)
    esperando_corte = False
    timer_activo = False
    estado_label.config(text="SISTEMA LISTO")

def recepcion_pulso():
    global esperando_corte
    if esperando_corte:
        return
    esperando_corte = True
    led.on() 
    canvas.itemconfig(indicador_led, fill=COLOR_OK)
    canvas.itemconfig(indicador_pulso, fill="#2196F3") 
    estado_label.config(text="Monitoreando carga...")

sensor_pulso.when_pressed = recepcion_pulso

def simular_pulso():
    """Botón extra para simular un pulso en la demo."""
    recepcion_pulso()

def reset_contadores():
    global piezas_ok, piezas_nok
    piezas_ok = 0
    piezas_nok = 0
    label_ok.config(text="OK: 0")
    label_nok.config(text="NOK: 0")
    label_ultimo_valor.config(text="--.-", fg=COLOR_TEXTO)
    estado_label.config(text="Contadores reiniciados")

# --- 6. REFRESCO DE INTERFAZ Y GRÁFICA ---
def refrescar_gui():
    global evento_corte_ui, tiempo_timer, limite_ok
    
    try: tiempo_timer = float(entry_timer.get())
    except ValueError: pass
    try: limite_ok = float(entry_limite_ok.get())
    except ValueError: pass

    # Evitar ruido en la gráfica: solo graficar si supera el umbral de activación
    val_to_plot = fuerza_actual if fuerza_actual > UMBRAL_ACTIVACION_N else 0.0
    datos_fuerza.pop(0)
    datos_fuerza.append(val_to_plot)
    datos_limite.pop(0)
    datos_limite.append(limite_ok)
    
    linea_fuerza.set_ydata(datos_fuerza)
    linea_limite.set_ydata(datos_limite)
    canvas_grafica.draw_idle() 

    # label_carga eliminado para evitar ruido en pantalla
    pass
        
    if evento_corte_ui:
        canvas.itemconfig(indicador_led, fill=COLOR_NOK)
        canvas.itemconfig(indicador_pulso, fill=COLOR_BORDE)
        label_ok.config(text=f"OK: {piezas_ok}")
        label_nok.config(text=f"NOK: {piezas_nok}")
        
        if tipo_corte == "OK":
            label_ultimo_valor.config(text=f"{valor_corte:.1f} N", fg=COLOR_OK)
            estado_label.config(text=f"¡Pieza OK! ({valor_corte:.1f} N)")
        else:
            label_ultimo_valor.config(text=f"{valor_corte:.1f} N", fg=COLOR_NOK)
            estado_label.config(text=f"¡Pieza NOK! ({valor_corte:.1f} N)")
            
        evento_corte_ui = False
        
    ventana.after(50, refrescar_gui)

def cerrar():
    global hilo_activo
    hilo_activo = False 
    led.off()
    sensor_pulso.close()
    ventana.destroy()


# ============================================================
# --- 7. DISEÑO DE INTERFAZ GRÁFICA (TEMA CLARO WKK) ---
# ============================================================

ventana = tk.Tk()
ventana.title("WKK - Sistema de Control de Prensa [DEMO]")
ventana.configure(bg=COLOR_FONDO)
ventana.geometry("1024x600")
ventana.minsize(900, 500)

# --- Helper: crear tarjeta con borde sutil ---
def crear_tarjeta(parent, **kwargs):
    return tk.Frame(parent, bg=COLOR_TARJETA, highlightbackground=COLOR_BORDE,
                    highlightthickness=1, padx=12, pady=8, **kwargs)

# ===================================================================
# HEADER — Logo WKK prominente + título
# ===================================================================
header = tk.Frame(ventana, bg=COLOR_TARJETA, height=65)
header.pack(fill="x", side="top")
header.pack_propagate(False)

# Cargar Logo WKK (grande y visible)
try:
    if PIL_DISPONIBLE:
        wkk_img = Image.open(os.path.join(ASSETS_DIR, "Logo WKK.png"))
        wkk_ratio = wkk_img.width / wkk_img.height
        wkk_h = 48
        wkk_w = int(wkk_h * wkk_ratio)
        wkk_img = wkk_img.resize((wkk_w, wkk_h), RESAMPLE_METHOD)
        wkk_photo = ImageTk.PhotoImage(wkk_img)
        wkk_label = tk.Label(header, image=wkk_photo, bg=COLOR_TARJETA)
        wkk_label.image = wkk_photo
        wkk_label.pack(side="left", padx=(18, 10), pady=8)
    else:
        raise ImportError("PIL no disponible")
except Exception as e:
    print(f"No se pudo cargar logo WKK: {e}")
    wkk_label = tk.Label(header, text="WKK", font=("Helvetica", 26, "bold"),
                         fg=COLOR_VERDE_WKK, bg=COLOR_TARJETA)
    wkk_label.pack(side="left", padx=(18, 10))

# Separador vertical decorativo
sep_v = tk.Frame(header, bg=COLOR_VERDE_WKK, width=3, height=40)
sep_v.pack(side="left", padx=(0, 15), pady=12)

titulo_header = tk.Label(header, text="Sistema de Control de Prensa",
                         font=("Helvetica", 18, "bold"), fg=COLOR_TEXTO, bg=COLOR_TARJETA)
titulo_header.pack(side="left", pady=10)

# Badge DEMO
demo_badge = tk.Label(header, text="  DEMO  ", font=("Helvetica", 9, "bold"),
                      fg="white", bg="#FF9800")
demo_badge.pack(side="left", padx=10, pady=20)

# Botón cerrar
btn_salir = tk.Button(header, text="✕", font=("Helvetica", 14, "bold"),
                      fg=COLOR_TEXTO_SEC, bg=COLOR_TARJETA, bd=0, padx=12,
                      activebackground=COLOR_NOK, activeforeground="white",
                      command=cerrar)
btn_salir.pack(side="right", padx=15)

# Línea de acento verde bajo el header
accent_top = tk.Frame(ventana, bg=COLOR_VERDE_WKK, height=3)
accent_top.pack(fill="x", side="top")

# ===================================================================
# FOOTER — Logo Baluarte pequeño + estado del sistema
# ===================================================================
footer = tk.Frame(ventana, bg=COLOR_TARJETA, height=38)
footer.pack(fill="x", side="bottom")
footer.pack_propagate(False)

# Línea de acento verde sobre el footer
accent_bottom = tk.Frame(ventana, bg=COLOR_VERDE_WKK, height=2)
accent_bottom.pack(fill="x", side="bottom")

# Logo Baluarte (pequeño, crédito)
try:
    if PIL_DISPONIBLE:
        bal_img = Image.open(os.path.join(ASSETS_DIR, "Logo Horizontal sin fondo.png"))
        bal_ratio = bal_img.width / bal_img.height
        bal_h = 24
        bal_w = int(bal_h * bal_ratio)
        bal_img = bal_img.resize((bal_w, bal_h), RESAMPLE_METHOD)
        bal_photo = ImageTk.PhotoImage(bal_img)
        bal_label = tk.Label(footer, image=bal_photo, bg=COLOR_TARJETA)
        bal_label.image = bal_photo
        bal_label.pack(side="left", padx=12, pady=6)
    else:
        raise ImportError("PIL no disponible")
except Exception as e:
    print(f"No se pudo cargar logo Baluarte: {e}")
    bal_label = tk.Label(footer, text="Baluarte", font=("Helvetica", 8, "italic"),
                         fg=COLOR_TEXTO_SEC, bg=COLOR_TARJETA)
    bal_label.pack(side="left", padx=12)

# Estado del sistema
estado_label = tk.Label(footer, text="SISTEMA LISTO", font=("Helvetica", 11, "bold"),
                        fg=COLOR_VERDE_WKK, bg=COLOR_TARJETA)
estado_label.pack(side="right", padx=18, pady=6)

# ===================================================================
# ÁREA DE CONTENIDO PRINCIPAL
# ===================================================================
content = tk.Frame(ventana, bg=COLOR_FONDO)
content.pack(fill="both", expand=True, padx=8, pady=6)

# --- Panel Izquierdo (Controles) ---
frame_izquierdo = tk.Frame(content, bg=COLOR_FONDO, width=360)
frame_izquierdo.pack(side="left", fill="y", padx=(0, 6))
frame_izquierdo.pack_propagate(False)

# --- Panel Derecho (Gráfica) ---
frame_derecho = tk.Frame(content, bg=COLOR_FONDO)
frame_derecho.pack(side="right", fill="both", expand=True)

# ===================================================================
# PANEL IZQUIERDO — Tarjetas de control
# ===================================================================

# (Tarjeta 1: Display de Fuerza eliminado para evitar ruido en pantalla)

# ─── Tarjeta 2: Parámetros ────────────────────────────────────────
card_params = crear_tarjeta(frame_izquierdo)
card_params.pack(fill="x", pady=(0, 5))

tk.Label(card_params, text="PARÁMETROS", font=("Helvetica", 9, "bold"),
         fg=COLOR_VERDE_WKK, bg=COLOR_TARJETA).pack(pady=(0, 4))

frame_params_grid = tk.Frame(card_params, bg=COLOR_TARJETA)
frame_params_grid.pack(fill="x")

# Tiempo Timer
tk.Label(frame_params_grid, text="Tiempo Timer (s):", font=("Helvetica", 10),
         fg=COLOR_TEXTO, bg=COLOR_TARJETA).grid(row=0, column=0, sticky="e", pady=2, padx=(0, 8))
entry_timer = tk.Entry(frame_params_grid, font=("Helvetica", 12), width=8, justify="center",
                        bg="#F8F9FA", fg=COLOR_TEXTO, insertbackground=COLOR_TEXTO,
                        highlightbackground=COLOR_VERDE_WKK, highlightthickness=1, relief="flat", bd=2)
entry_timer.insert(0, "2.0")
entry_timer.grid(row=0, column=1, pady=2, sticky="w")

# Límite OK (N)
tk.Label(frame_params_grid, text="Fuerza Mínima OK (N):", font=("Helvetica", 10),
         fg=COLOR_TEXTO, bg=COLOR_TARJETA).grid(row=1, column=0, sticky="e", pady=2, padx=(0, 8))
entry_limite_ok = tk.Entry(frame_params_grid, font=("Helvetica", 12), width=8, justify="center",
                     bg="#F8F9FA", fg=COLOR_TEXTO, insertbackground=COLOR_TEXTO,
                     highlightbackground=COLOR_VERDE_WKK, highlightthickness=1, relief="flat", bd=2)
entry_limite_ok.insert(0, "598.2")
entry_limite_ok.grid(row=1, column=1, pady=2, sticky="w")

# ─── Tarjeta 3: Calidad ───────────────────────────────────────────
card_calidad = crear_tarjeta(frame_izquierdo)
card_calidad.pack(fill="x", pady=(0, 5))

tk.Label(card_calidad, text="CALIDAD", font=("Helvetica", 9, "bold"),
         fg=COLOR_VERDE_WKK, bg=COLOR_TARJETA).pack(pady=(0, 4))

frame_ultimo = tk.Frame(card_calidad, bg=COLOR_TARJETA)
frame_ultimo.pack(fill="x")
tk.Label(frame_ultimo, text="Último valor:", font=("Helvetica", 10),
         fg=COLOR_TEXTO_SEC, bg=COLOR_TARJETA).pack(side="left", padx=(0, 8))
label_ultimo_valor = tk.Label(frame_ultimo, text="--.-", font=("Helvetica", 15, "bold"),
                              fg=COLOR_TEXTO, bg=COLOR_TARJETA)
label_ultimo_valor.pack(side="left")

frame_contadores = tk.Frame(card_calidad, bg=COLOR_TARJETA)
frame_contadores.pack(fill="x", pady=6)

label_ok = tk.Label(frame_contadores, text="OK: 0", font=("Helvetica", 12, "bold"),
                    fg="white", bg=COLOR_OK, pady=5)
label_ok.pack(side="left", padx=(0, 6), expand=True, fill="x")

label_nok = tk.Label(frame_contadores, text="NOK: 0", font=("Helvetica", 12, "bold"),
                     fg="white", bg=COLOR_NOK, pady=5)
label_nok.pack(side="left", expand=True, fill="x")

btn_reset = tk.Button(card_calidad, text="↺  Reset Contadores", font=("Helvetica", 9),
                      fg=COLOR_TEXTO_SEC, bg="#F1F3F5", bd=0, pady=4,
                      activebackground=COLOR_BORDE, command=reset_contadores)
btn_reset.pack(fill="x", pady=(3, 0))

# ─── Tarjeta 4: Indicadores y Controles ───────────────────────────
card_controles = crear_tarjeta(frame_izquierdo)
card_controles.pack(fill="x", pady=(0, 5))

tk.Label(card_controles, text="CONTROLES", font=("Helvetica", 9, "bold"),
         fg=COLOR_VERDE_WKK, bg=COLOR_TARJETA).pack(pady=(0, 3))

frame_indicadores = tk.Frame(card_controles, bg=COLOR_TARJETA)
frame_indicadores.pack(fill="x", pady=2)

canvas = tk.Canvas(frame_indicadores, width=220, height=45, bg=COLOR_TARJETA,
                   highlightthickness=0)
canvas.pack()

canvas.create_text(55, 8, text="Pulso (Pin 4)", font=("Helvetica", 8, "bold"), fill=COLOR_TEXTO_SEC)
indicador_pulso = canvas.create_oval(22, 16, 88, 43, fill=COLOR_BORDE, outline=COLOR_TEXTO_SEC, width=1)

canvas.create_text(165, 8, text="Estado LED", font=("Helvetica", 8, "bold"), fill=COLOR_TEXTO_SEC)
indicador_led = canvas.create_oval(132, 16, 198, 43, fill=COLOR_NOK, outline=COLOR_TEXTO_SEC, width=1)

frame_botones = tk.Frame(card_controles, bg=COLOR_TARJETA)
frame_botones.pack(fill="x", pady=4)

btn_encender = tk.Button(frame_botones, text="● Encender", font=("Helvetica", 10, "bold"),
                         fg="white", bg=COLOR_OK, bd=0, pady=7,
                         activebackground="#00A844", command=encender_led_manual)
btn_encender.pack(side="left", expand=True, fill="x", padx=(0, 4))

btn_apagar = tk.Button(frame_botones, text="■ Apagar", font=("Helvetica", 10, "bold"),
                       fg="white", bg=COLOR_NOK, bd=0, pady=7,
                       activebackground="#C62828", command=apagar_led_manual)
btn_apagar.pack(side="left", expand=True, fill="x", padx=(4, 0))

# Botón Simular Pulso (SOLO EN DEMO)
btn_simular = tk.Button(card_controles, text="⚡  Simular Pulso (Demo)", font=("Helvetica", 9, "bold"),
                        fg="white", bg="#FF9800", bd=0, pady=6,
                        activebackground="#F57C00", command=simular_pulso)
btn_simular.pack(fill="x", pady=(4, 0))

# Botón Exportar Excel
btn_exportar = tk.Button(card_controles, text="📊  Exportar a Excel", font=("Helvetica", 10, "bold"),
                         fg="white", bg=COLOR_VERDE_WKK, bd=0, pady=8,
                         activebackground=COLOR_VERDE_OSCURO, command=exportar_a_excel)
btn_exportar.pack(fill="x", pady=(4, 0))

# ===================================================================
# PANEL DERECHO — Gráfica en tiempo real
# ===================================================================
card_grafica = crear_tarjeta(frame_derecho)
card_grafica.pack(fill="both", expand=True)

tk.Label(card_grafica, text="COMPORTAMIENTO DE FUERZA VS SETPOINTS",
         font=("Helvetica", 9, "bold"), fg=COLOR_VERDE_WKK, bg=COLOR_TARJETA).pack(pady=(0, 4))

figura = Figure(dpi=85)
figura.patch.set_facecolor(COLOR_TARJETA)
ax = figura.add_subplot(111)
ax.set_facecolor("#FAFAFA")
ax.set_xlabel("Muestras en Tiempo Real", fontsize=9, color=COLOR_TEXTO_SEC)
ax.set_ylabel("Fuerza / Presión", fontsize=9, color=COLOR_TEXTO_SEC)
ax.set_ylim(-2, 105) 
ax.grid(True, linestyle="--", alpha=0.3, color="#ADB5BD")

ax.spines['top'].set_visible(False)
ax.spines['right'].set_visible(False)
ax.spines['left'].set_color(COLOR_BORDE)
ax.spines['bottom'].set_color(COLOR_BORDE)
ax.tick_params(colors=COLOR_TEXTO_SEC, labelsize=8)

linea_fuerza, = ax.plot(datos_fuerza, label="Fuerza Actual (N)", color=COLOR_VERDE_WKK, linewidth=2.5)
linea_limite, = ax.plot(datos_limite, label="Límite Mínimo OK (N)", color="#2196F3", linestyle="--", linewidth=1.5)
ax.legend(loc="upper left", fontsize=8, framealpha=0.9, edgecolor=COLOR_BORDE)
figura.tight_layout(pad=2)

canvas_grafica = FigureCanvasTkAgg(figura, master=card_grafica)
canvas_grafica.get_tk_widget().pack(fill="both", expand=True)

# --- 8. INICIAR SISTEMA ---
ventana.protocol("WM_DELETE_WINDOW", cerrar)

hilo_modbus = threading.Thread(target=tarea_modbus_alta_velocidad, daemon=True)
hilo_modbus.start()

refrescar_gui() 
ventana.mainloop()
