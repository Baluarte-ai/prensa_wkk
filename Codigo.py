import tkinter as tk
from tkinter import messagebox, filedialog, ttk
import minimalmodbus
from gpiozero import LED, Button
import threading
import time
import sqlite3
import os
from datetime import datetime

# --- IMPORTACIONES PARA LA GRÁFICA ---
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure

# --- IMPORTACIONES PARA LOGOS ---
try:
    from PIL import Image, ImageTk
    PIL_DISPONIBLE = True
    try:
        RESAMPLE_METHOD = Image.LANCZOS
    except AttributeError:
        RESAMPLE_METHOD = Image.ANTIALIAS
except ImportError:
    PIL_DISPONIBLE = False

# --- PALETA DE COLORES WKK ---
COLOR_VERDE_WKK = "#0E8A3E"
COLOR_VERDE_OSCURO = "#0A6B2F"
COLOR_VERDE_CLARO = "#E8F5E9"
COLOR_FONDO = "#F0F2F5"
COLOR_TARJETA = "#FFFFFF"
COLOR_BORDE = "#DEE2E6"
COLOR_TEXTO = "#1A1A2E"
COLOR_TEXTO_SEC = "#6C757D"
COLOR_OK = "#00C853"
COLOR_NOK = "#E53935"

# --- 1. CONFIGURACIÓN DE HARDWARE (GPIO) ---
led = LED(17, initial_value=True)  # initial_value=True = GPIO HIGH al inicio (relay inactivo)
sensor_pulso = Button(4, pull_up=False)

# Entradas de seguridad (Pin 26 y Pin 22, activo-alto / pull_up=False)
try:
    barrera = Button(26, pull_up=False)
    barrera_conectada = True
except Exception:
    barrera_conectada = False

try:
    paro_emergencia = Button(22, pull_up=False)
    paro_conectado = True
except Exception:
    paro_conectado = False

# --- VARIABLES GLOBALES DE CONTROL ---
hilo_activo = True
fuerza_actual = 0.0
esperando_corte = False
evento_corte_ui = False
tipo_corte = ""
valor_corte = 0.0
piezas_ok = 0
piezas_nok = 0

umbral_global = 25.0
min_global = 61.4
max_global = 64.4

# --- VARIABLES PARA EL HISTORIAL DE LA GRÁFICA ---
MAX_PUNTOS = 100  
datos_fuerza = [0.0] * MAX_PUNTOS
datos_umbral = [25.0] * MAX_PUNTOS
datos_min = [61.4] * MAX_PUNTOS
datos_max = [64.4] * MAX_PUNTOS

# --- 2. CONFIGURACIÓN DE MODBUS (Velocidad Equilibrada) ---
try:
    instrument = minimalmodbus.Instrument('/dev/ttyUSB0', 1)
    
    # Regresamos a 9600 para garantizar que el sensor físico y la placa se entiendan
    instrument.serial.baudrate = 9600 
    
    instrument.serial.bytesize = 8
    instrument.serial.parity = minimalmodbus.serial.PARITY_NONE
    instrument.serial.stopbits = 1
    
    # Timeout intermedio: 30 milisegundos. Suficiente para recibir el dato sin congelarse.
    instrument.serial.timeout = 0.03 
    modbus_conectado = True
except Exception as e:
    print(f"Advertencia: No se pudo conectar Modbus. {e}")
    modbus_conectado = False

# --- 3. CALIBRACIÓN MATEMÁTICA ---
LECTURA_CERO = 7.43     
LECTURA_MAXIMA = 8.23   
FUERZA_MAXIMA = 100.0   

# --- RUTAS ---
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ASSETS_DIR = os.path.join(SCRIPT_DIR, "assets")
DB_PATH = os.path.join(SCRIPT_DIR, "registros_prensa.db")

# --- BASE DE DATOS SQLite ---
def inicializar_db():
    """Crea la tabla de registros si no existe."""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS registros (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha TEXT NOT NULL,
            hora TEXT NOT NULL,
            valor_fuerza REAL NOT NULL,
            resultado TEXT NOT NULL,
            umbral REAL,
            minimo_ok REAL,
            maximo_ok REAL
        )
    """)
    conn.commit()
    conn.close()

inicializar_db()

def registrar_corte(valor, resultado, umbral, minimo, maximo):
    """Registra un evento de corte en la base de datos."""
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        ahora = datetime.now()
        cursor.execute(
            "INSERT INTO registros (fecha, hora, valor_fuerza, resultado, umbral, minimo_ok, maximo_ok) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (ahora.strftime("%Y-%m-%d"), ahora.strftime("%H:%M:%S"), valor, resultado, umbral, minimo, maximo)
        )
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"Error al registrar en DB: {e}")

def abrir_previsualizacion():
    """Muestra una ventana modal con los registros actuales antes de exportar."""
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("SELECT id, fecha, hora, valor_fuerza, resultado, umbral, minimo_ok, maximo_ok FROM registros ORDER BY id DESC")
        registros = cursor.fetchall()
        conn.close()
    except Exception as e:
        messagebox.showerror("Error de Base de Datos", f"No se pudo leer la base de datos:\n{e}")
        return

    if not registros:
        messagebox.showinfo("Previsualización", "No hay registros en la base de datos para mostrar.")
        return

    # Desactivar fullscreen temporalmente para mejor interacción modal en la tablet
    ventana.attributes('-fullscreen', False)
    ventana.update()

    preview_win = tk.Toplevel(ventana)
    preview_win.title("WKK - Previsualización de Registros")
    preview_win.configure(bg=COLOR_FONDO)
    
    # Tamaño y centrado para la tablet 1900x1200
    w_width, w_height = 950, 650
    screen_w = preview_win.winfo_screenwidth()
    screen_h = preview_win.winfo_screenheight()
    pos_x = (screen_w - w_width) // 2
    pos_y = (screen_h - w_height) // 2
    preview_win.geometry(f"{w_width}x{w_height}+{pos_x}+{pos_y}")
    preview_win.transient(ventana)
    preview_win.grab_set()

    # Título del modal
    header_frame = tk.Frame(preview_win, bg=COLOR_TARJETA, pady=12)
    header_frame.pack(fill="x")
    tk.Label(header_frame, text="REGISTROS DE CORTE ALMACENADOS", font=("Helvetica", 14, "bold"),
             fg=COLOR_VERDE_WKK, bg=COLOR_TARJETA).pack()

    # Contenedor para tabla y scrollbar
    table_frame = tk.Frame(preview_win, bg=COLOR_FONDO, padx=15, pady=15)
    table_frame.pack(fill="both", expand=True)

    # Configuración de estilos para el Treeview (tabla)
    style = ttk.Style()
    style.theme_use("clam")
    style.configure("Treeview",
                    background=COLOR_TARJETA,
                    foreground=COLOR_TEXTO,
                    rowheight=35,
                    fieldbackground=COLOR_TARJETA,
                    font=("Helvetica", 11))
    style.configure("Treeview.Heading",
                    background=COLOR_VERDE_WKK,
                    foreground="white",
                    font=("Helvetica", 11, "bold"))
    style.map("Treeview.Heading", background=[('active', COLOR_VERDE_OSCURO)])

    columns = ("id", "fecha", "hora", "fuerza", "resultado", "umbral", "min", "max")
    tree = ttk.Treeview(table_frame, columns=columns, show="headings", style="Treeview")
    
    # Definir encabezados
    headers = {
        "id": "ID", "fecha": "Fecha", "hora": "Hora",
        "fuerza": "Fuerza (kg)", "resultado": "Resultado",
        "umbral": "Umbral", "min": "Mínimo OK", "max": "Máximo OK"
    }
    for col, txt in headers.items():
        tree.heading(col, text=txt)
        tree.column(col, anchor="center", width=100)
    tree.column("id", width=60)
    tree.column("resultado", width=120)

    # Scrollbar
    scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=scrollbar.set)
    
    tree.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")

    # Configurar tags de color para las filas OK y NOK
    tree.tag_configure("OK", background="#E8F5E9", foreground="#2E7D32")
    tree.tag_configure("NOK", background="#FFEBEE", foreground="#C62828")

    # Insertar datos
    for reg in registros:
        res = reg[4] # "OK" o "NOK"
        tree.insert("", "end", values=reg, tags=(res,))

    # Botonera inferior
    btn_frame = tk.Frame(preview_win, bg=COLOR_TARJETA, pady=15)
    btn_frame.pack(fill="x", side="bottom")

    def ejecutar_exportado():
        preview_win.destroy()
        exportar_a_excel()

    def cancelar_modal():
        preview_win.destroy()
        ventana.attributes('-fullscreen', True)

    btn_guardar = tk.Button(btn_frame, text="📊  Exportar a Excel", font=("Helvetica", 12, "bold"),
                            fg="white", bg=COLOR_VERDE_WKK, bd=0, padx=20, pady=10,
                            activebackground=COLOR_VERDE_OSCURO, command=ejecutar_exportado)
    btn_guardar.pack(side="left", padx=(40, 20))

    btn_cancelar = tk.Button(btn_frame, text="✕  Cerrar", font=("Helvetica", 12, "bold"),
                             fg=COLOR_TEXTO_SEC, bg="#F1F3F5", bd=0, padx=20, pady=10,
                             activebackground=COLOR_BORDE, command=cancelar_modal)
    btn_cancelar.pack(side="right", padx=(20, 40))

    preview_win.protocol("WM_DELETE_WINDOW", cancelar_modal)

def exportar_a_excel():
    """Exporta todos los registros de la base de datos a un archivo Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM registros ORDER BY id")
        registros = cursor.fetchall()
        conn.close()
        
        if not registros:
            messagebox.showinfo("Exportar Excel", "No hay registros para exportar.")
            return
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Registros de Prensa"
        
        # Estilos del Excel
        header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='0E8A3E', end_color='0E8A3E', fill_type='solid')
        ok_fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
        nok_fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
        center = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Encabezados
        headers = ['ID', 'Fecha', 'Hora', 'Valor Fuerza', 'Resultado', 'Umbral', 'Mínimo OK', 'Máximo OK']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border
        
        # Datos
        for row_idx, registro in enumerate(registros, 2):
            for col_idx, val in enumerate(registro, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = center
                cell.border = thin_border
                if col_idx == 5:  # Columna resultado
                    cell.fill = ok_fill if val == "OK" else nok_fill
        
        # Ajustar ancho de columnas
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 15
        
        # Desactivar fullscreen temporalmente para que el diálogo sea visible en la tablet
        ventana.attributes('-fullscreen', False)
        ventana.update()
        
        # Detectar si hay una USB conectada en la Raspberry Pi
        initial_dir = SCRIPT_DIR
        media_user_path = "/media/baluarte_admin1"
        try:
            if os.path.exists(media_user_path):
                subdirs = [os.path.join(media_user_path, d) for d in os.listdir(media_user_path) 
                           if os.path.isdir(os.path.join(media_user_path, d))]
                if subdirs:
                    initial_dir = subdirs[0] # Abre directamente en la primera USB encontrada
        except Exception:
            pass
            
        # Diálogo "Guardar como" (detecta unidades externas/USB)
        fecha_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = filedialog.asksaveasfilename(
            parent=ventana,
            title="Guardar registros como...",
            initialdir=initial_dir,
            initialfile=f"registros_prensa_{fecha_str}.xlsx",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("Todos los archivos", "*.*")]
        )
        
        # Reactivar fullscreen
        ventana.attributes('-fullscreen', True)
        
        if not filename:  # El usuario canceló
            return
            
        try:
            wb.save(filename)
            messagebox.showinfo("Exportar Excel", f"Archivo exportado exitosamente:\n{filename}")
        except PermissionError:
            # Respaldar localmente si la USB es de solo lectura
            respaldo_local = os.path.join(SCRIPT_DIR, os.path.basename(filename))
            try:
                wb.save(respaldo_local)
                messagebox.showwarning(
                    "Error de Permisos",
                    f"No se pudo guardar en la ruta seleccionada (posiblemente la USB está protegida o es de solo lectura).\n\n"
                    f"Se guardó una copia de seguridad local en:\n{respaldo_local}"
                )
            except Exception as e_inner:
                messagebox.showerror("Error Crítico", f"No se pudo guardar ni en la USB ni localmente:\n{e_inner}")
        except Exception as e:
            messagebox.showerror("Error", f"Error al guardar el archivo:\n{e}")
            
    except ImportError:
        messagebox.showerror("Error", "Se requiere instalar openpyxl:\npip install openpyxl")
    except Exception as e:
        messagebox.showerror("Error", f"Error al exportar:\n{e}")

# --- 4. HILO DE LECTURA (HARDWARE Y MODBUS) ---
def tarea_modbus_alta_velocidad():
    global fuerza_actual, esperando_corte, evento_corte_ui
    global tipo_corte, valor_corte, piezas_ok, piezas_nok
    
    while hilo_activo:
        if not modbus_conectado:
            time.sleep(1)
            continue
            
        try:
            instrument.serial.reset_input_buffer() 
            lectura_bruta = instrument.read_register(0, 2, functioncode=3)
            
            f_calc = ((lectura_bruta - LECTURA_CERO) / (LECTURA_MAXIMA - LECTURA_CERO)) * FUERZA_MAXIMA
            
            if f_calc < 0.1: f_calc = 0.0
            if f_calc > FUERZA_MAXIMA: f_calc = FUERZA_MAXIMA
            
            fuerza_actual = f_calc
              # --- EVALUACIÓN A VELOCIDAD DE HARDWARE ---
            if esperando_corte:
                # Protección redundante en hilo de alta velocidad
                if (barrera_conectada and barrera.is_pressed) or (paro_conectado and paro_emergencia.is_pressed):
                    led.on() # Apagar electroválvula inmediatamente
                    esperando_corte = False
                    continue

                if fuerza_actual >= umbral_global:
                    led.on() # Corte de la electroválvula
                    
                    valor_corte = fuerza_actual
                    if min_global <= fuerza_actual <= max_global:
                        piezas_ok += 1
                        tipo_corte = "OK"
                    else:
                        piezas_nok += 1
                        tipo_corte = "NOK"
                    
                    # Registrar en base de datos
                    registrar_corte(valor_corte, tipo_corte, umbral_global, min_global, max_global)
                        
                    esperando_corte = False
                    evento_corte_ui = True
                    
        except Exception as e:
            # Si hay un pequeño error de lectura, simplemente lo ignora y sigue intentando
            pass 
            
        # Respiro de 5 milisegundos. Evita saturar el puerto USB y estabiliza la lectura
        time.sleep(0.005)

# --- 5. FUNCIONES DE INTERFAZ Y PULSO ---
def detener_por_seguridad():
    global esperando_corte
    led.on()  # Apagar pistón (inactivo)
    esperando_corte = False
    canvas.itemconfig(indicador_led, fill=COLOR_NOK)
    canvas.itemconfig(indicador_pulso, fill=COLOR_BORDE)
    estado_label.config(text="DETENIDO POR SEGURIDAD", fg=COLOR_NOK)

def encender_led_manual():
    # Evitar acción si la seguridad está activa
    if (barrera_conectada and barrera.is_pressed) or (paro_conectado and paro_emergencia.is_pressed):
        estado_label.config(text="BLOQUEADO: Seguridad Activa", fg=COLOR_NOK)
        return
    led.off()
    canvas.itemconfig(indicador_led, fill=COLOR_OK)
    estado_label.config(text="LED ENCENDIDO MANUAL")

def apagar_led_manual():
    global esperando_corte
    led.on()
    canvas.itemconfig(indicador_led, fill=COLOR_NOK)
    canvas.itemconfig(indicador_pulso, fill=COLOR_BORDE)
    esperando_corte = False
    estado_label.config(text="SISTEMA LISTO")

def recepcion_pulso():
    global esperando_corte
    if esperando_corte:
        return
    # Evitar acción si la seguridad está activa
    if (barrera_conectada and barrera.is_pressed) or (paro_conectado and paro_emergencia.is_pressed):
        estado_label.config(text="BLOQUEADO: Seguridad Activa", fg=COLOR_NOK)
        return
    esperando_corte = True
    led.off() 
    canvas.itemconfig(indicador_led, fill=COLOR_OK)
    canvas.itemconfig(indicador_pulso, fill="#2196F3") 
    estado_label.config(text="Monitoreando carga...")

# NOTA: El callback del sensor se registra en la sección 8,
# DESPUÉS de construir la GUI, para evitar pulsos espurios al iniciar.

def reset_contadores():
    global piezas_ok, piezas_nok
    piezas_ok = 0
    piezas_nok = 0
    label_ok.config(text="OK: 0")
    label_nok.config(text="NOK: 0")
    label_ultimo_valor.config(text="--.-", fg=COLOR_TEXTO)
    estado_label.config(text="Contadores reiniciados")

# --- 6. REFRESCO DE INTERFAZ Y GRÁFICA ---
def refrescar_gui():
    global evento_corte_ui, umbral_global, min_global, max_global
    
    # 1. Leer los Entry
    try: umbral_global = float(entry_umbral.get())
    except ValueError: pass
    try: min_global = float(entry_min.get())
    except ValueError: pass
    try: max_global = float(entry_max.get())
    except ValueError: pass

    # 2. Actualizar datos de la gráfica
    datos_fuerza.pop(0)
    datos_fuerza.append(fuerza_actual)
    datos_umbral.pop(0)
    datos_umbral.append(umbral_global)
    datos_min.pop(0)
    datos_min.append(min_global)
    datos_max.pop(0)
    datos_max.append(max_global)
    
    # 3. Dibujar gráfica
    linea_fuerza.set_ydata(datos_fuerza)
    linea_umbral.set_ydata(datos_umbral)
    linea_min.set_ydata(datos_min)
    linea_max.set_ydata(datos_max)
    canvas_grafica.draw_idle() 

    # 4. Actualizar textos
    if modbus_conectado:
        label_carga.config(text=f"{fuerza_actual:.1f}")
    else:
        label_carga.config(text="Error", fg=COLOR_NOK)

    # 5. Actualizar indicadores de seguridad (Activo-Alto: Pressed = Red/Warning)
    if barrera_conectada:
        if barrera.is_pressed:
            canvas.itemconfig(indicador_barrera, fill=COLOR_NOK)
            lbl_barrera_estado.config(text="¡ACTIVADO!", fg=COLOR_NOK)
        else:
            canvas.itemconfig(indicador_barrera, fill=COLOR_OK)
            lbl_barrera_estado.config(text="NORMAL", fg=COLOR_OK)
    
    if paro_conectado:
        if paro_emergencia.is_pressed:
            canvas.itemconfig(indicador_paro, fill=COLOR_NOK)
            lbl_paro_estado.config(text="¡ACTIVADO!", fg=COLOR_NOK)
        else:
            canvas.itemconfig(indicador_paro, fill=COLOR_OK)
            lbl_paro_estado.config(text="NORMAL", fg=COLOR_OK)
        
    # 6. Respuesta visual de corte
    if evento_corte_ui:
        canvas.itemconfig(indicador_led, fill=COLOR_NOK)
        canvas.itemconfig(indicador_pulso, fill=COLOR_BORDE)
        label_ok.config(text=f"OK: {piezas_ok}")
        label_nok.config(text=f"NOK: {piezas_nok}")
        
        if tipo_corte == "OK":
            label_ultimo_valor.config(text=f"{valor_corte:.1f}", fg=COLOR_OK)
            estado_label.config(text=f"¡Pieza OK! ({valor_corte:.1f})")
        else:
            label_ultimo_valor.config(text=f"{valor_corte:.1f}", fg=COLOR_NOK)
            estado_label.config(text=f"¡Pieza NOK! ({valor_corte:.1f})")
            
        evento_corte_ui = False
        
    ventana.after(50, refrescar_gui)

def cerrar():
    global hilo_activo
    hilo_activo = False 
    led.on()  # GPIO HIGH = relay inactivo (circuito activo-bajo)
    sensor_pulso.close()
    if barrera_conectada: barrera.close()
    if paro_conectado: paro_emergencia.close()
    ventana.destroy()


# ============================================================
# --- 7. DISEÑO DE INTERFAZ GRÁFICA (TEMA CLARO WKK) ---
# ============================================================

ventana = tk.Tk()
ventana.title("WKK - Sistema de Control de Prensa")
ventana.configure(bg=COLOR_FONDO)
ventana.attributes('-fullscreen', True)
ventana.bind('<Escape>', lambda e: cerrar())

# --- Helper: crear tarjeta con borde sutil ---
def crear_tarjeta(parent, **kwargs):
    return tk.Frame(parent, bg=COLOR_TARJETA, highlightbackground=COLOR_BORDE,
                    highlightthickness=1, padx=18, pady=12, **kwargs)

# ===================================================================
# HEADER — Logo WKK prominente + título
# ===================================================================
header = tk.Frame(ventana, bg=COLOR_TARJETA, height=85)
header.pack(fill="x", side="top")
header.pack_propagate(False)

# Cargar Logo WKK (grande y visible)
try:
    if PIL_DISPONIBLE:
        wkk_img = Image.open(os.path.join(ASSETS_DIR, "Logo WKK.png"))
        wkk_ratio = wkk_img.width / wkk_img.height
        wkk_h = 62
        wkk_w = int(wkk_h * wkk_ratio)
        wkk_img = wkk_img.resize((wkk_w, wkk_h), RESAMPLE_METHOD)
        wkk_photo = ImageTk.PhotoImage(wkk_img)
        wkk_label = tk.Label(header, image=wkk_photo, bg=COLOR_TARJETA)
        wkk_label.image = wkk_photo  # Mantener referencia
        wkk_label.pack(side="left", padx=(18, 10), pady=8)
    else:
        raise ImportError("PIL no disponible")
except Exception as e:
    print(f"No se pudo cargar logo WKK: {e}")
    wkk_label = tk.Label(header, text="WKK", font=("Helvetica", 34, "bold"),
                         fg=COLOR_VERDE_WKK, bg=COLOR_TARJETA)
    wkk_label.pack(side="left", padx=(18, 10))

# Separador vertical decorativo
sep_v = tk.Frame(header, bg=COLOR_VERDE_WKK, width=3, height=55)
sep_v.pack(side="left", padx=(0, 20), pady=15)

titulo_header = tk.Label(header, text="Sistema de Control de Prensa",
                         font=("Helvetica", 24, "bold"), fg=COLOR_TEXTO, bg=COLOR_TARJETA)
titulo_header.pack(side="left", pady=15)

# Botón cerrar
btn_salir = tk.Button(header, text="✕", font=("Helvetica", 18, "bold"),
                      fg=COLOR_TEXTO_SEC, bg=COLOR_TARJETA, bd=0, padx=18, pady=8,
                      activebackground=COLOR_NOK, activeforeground="white",
                      command=cerrar)
btn_salir.pack(side="right", padx=20)

# Línea de acento verde bajo el header
accent_top = tk.Frame(ventana, bg=COLOR_VERDE_WKK, height=4)
accent_top.pack(fill="x", side="top")

# ===================================================================
# FOOTER — Logo Baluarte pequeño + estado del sistema
# ===================================================================
footer = tk.Frame(ventana, bg=COLOR_TARJETA, height=50)
footer.pack(fill="x", side="bottom")
footer.pack_propagate(False)

# Línea de acento verde sobre el footer
accent_bottom = tk.Frame(ventana, bg=COLOR_VERDE_WKK, height=2)
accent_bottom.pack(fill="x", side="bottom")

# Logo Baluarte (pequeño, crédito)
try:
    if PIL_DISPONIBLE:
        bal_img = Image.open(os.path.join(ASSETS_DIR, "Logo Horizontal sin fondo.png"))
        bal_ratio = bal_img.width / bal_img.height
        bal_h = 32
        bal_w = int(bal_h * bal_ratio)
        bal_img = bal_img.resize((bal_w, bal_h), RESAMPLE_METHOD)
        bal_photo = ImageTk.PhotoImage(bal_img)
        bal_label = tk.Label(footer, image=bal_photo, bg=COLOR_TARJETA)
        bal_label.image = bal_photo
        bal_label.pack(side="left", padx=12, pady=6)
    else:
        raise ImportError("PIL no disponible")
except Exception as e:
    print(f"No se pudo cargar logo Baluarte: {e}")
    bal_label = tk.Label(footer, text="Baluarte", font=("Helvetica", 8, "italic"),
                         fg=COLOR_TEXTO_SEC, bg=COLOR_TARJETA)
    bal_label.pack(side="left", padx=12)

# Estado del sistema (en el footer)
estado_label = tk.Label(footer, text="SISTEMA LISTO", font=("Helvetica", 14, "bold"),
                        fg=COLOR_VERDE_WKK, bg=COLOR_TARJETA)
estado_label.pack(side="right", padx=24, pady=10)

# ===================================================================
# ÁREA DE CONTENIDO PRINCIPAL
# ===================================================================
content = tk.Frame(ventana, bg=COLOR_FONDO)
content.pack(fill="both", expand=True, padx=14, pady=10)

# --- Panel Izquierdo (Controles) ---
frame_izquierdo = tk.Frame(content, bg=COLOR_FONDO, width=520)
frame_izquierdo.pack(side="left", fill="y", padx=(0, 10))
frame_izquierdo.pack_propagate(False)

# --- Panel Derecho (Gráfica) ---
frame_derecho = tk.Frame(content, bg=COLOR_FONDO)
frame_derecho.pack(side="right", fill="both", expand=True)

# ===================================================================
# PANEL IZQUIERDO — Tarjetas de control
# ===================================================================

# ─── Tarjeta 1: Display de Fuerza ─────────────────────────────────
card_fuerza = crear_tarjeta(frame_izquierdo)
card_fuerza.pack(fill="x", pady=(0, 5))

lbl_titulo_fuerza = tk.Label(card_fuerza, text="FUERZA / PRESIÓN ACTUAL",
                             font=("Helvetica", 13, "bold"), fg=COLOR_VERDE_WKK, bg=COLOR_TARJETA)
lbl_titulo_fuerza.pack(pady=(0, 6))

frame_display = tk.Frame(card_fuerza, bg="#1B2838", padx=20, pady=10,
                         highlightbackground=COLOR_VERDE_WKK, highlightthickness=3)
frame_display.pack(fill="x")

label_carga = tk.Label(frame_display, text="0.0", font=("Helvetica", 64, "bold"),
                       fg="#00E676", bg="#1B2838")
label_carga.pack()

# ─── Tarjeta 2: Parámetros ────────────────────────────────────────
card_params = crear_tarjeta(frame_izquierdo)
card_params.pack(fill="x", pady=(0, 5))

tk.Label(card_params, text="PARÁMETROS", font=("Helvetica", 13, "bold"),
         fg=COLOR_VERDE_WKK, bg=COLOR_TARJETA).pack(pady=(0, 6))

frame_params_grid = tk.Frame(card_params, bg=COLOR_TARJETA)
frame_params_grid.pack(fill="x")

# Umbral
tk.Label(frame_params_grid, text="Umbral de Corte:", font=("Helvetica", 13),
         fg=COLOR_TEXTO, bg=COLOR_TARJETA).grid(row=0, column=0, sticky="e", pady=4, padx=(0, 12))
entry_umbral = tk.Entry(frame_params_grid, font=("Helvetica", 15), width=8, justify="center",
                        bg="#F8F9FA", fg=COLOR_TEXTO, insertbackground=COLOR_TEXTO,
                        highlightbackground=COLOR_VERDE_WKK, highlightthickness=1, relief="flat", bd=3)
entry_umbral.insert(0, "25.0")
entry_umbral.grid(row=0, column=1, pady=4, sticky="w")

# Mínimo
tk.Label(frame_params_grid, text="Mínimo OK:", font=("Helvetica", 13),
         fg=COLOR_TEXTO, bg=COLOR_TARJETA).grid(row=1, column=0, sticky="e", pady=4, padx=(0, 12))
entry_min = tk.Entry(frame_params_grid, font=("Helvetica", 15), width=8, justify="center",
                     bg="#F8F9FA", fg=COLOR_TEXTO, insertbackground=COLOR_TEXTO,
                     highlightbackground=COLOR_VERDE_WKK, highlightthickness=1, relief="flat", bd=3)
entry_min.insert(0, "61.4")
entry_min.grid(row=1, column=1, pady=4, sticky="w")

# Máximo
tk.Label(frame_params_grid, text="Máximo OK:", font=("Helvetica", 13),
         fg=COLOR_TEXTO, bg=COLOR_TARJETA).grid(row=2, column=0, sticky="e", pady=4, padx=(0, 12))
entry_max = tk.Entry(frame_params_grid, font=("Helvetica", 15), width=8, justify="center",
                     bg="#F8F9FA", fg=COLOR_TEXTO, insertbackground=COLOR_TEXTO,
                     highlightbackground=COLOR_VERDE_WKK, highlightthickness=1, relief="flat", bd=3)
entry_max.insert(0, "64.4")
entry_max.grid(row=2, column=1, pady=4, sticky="w")

# ─── Tarjeta 3: Calidad ───────────────────────────────────────────
card_calidad = crear_tarjeta(frame_izquierdo)
card_calidad.pack(fill="x", pady=(0, 5))

tk.Label(card_calidad, text="CALIDAD", font=("Helvetica", 13, "bold"),
         fg=COLOR_VERDE_WKK, bg=COLOR_TARJETA).pack(pady=(0, 6))

# Último valor registrado
frame_ultimo = tk.Frame(card_calidad, bg=COLOR_TARJETA)
frame_ultimo.pack(fill="x")
tk.Label(frame_ultimo, text="Último valor:", font=("Helvetica", 13),
         fg=COLOR_TEXTO_SEC, bg=COLOR_TARJETA).pack(side="left", padx=(0, 10))
label_ultimo_valor = tk.Label(frame_ultimo, text="--.-", font=("Helvetica", 20, "bold"),
                              fg=COLOR_TEXTO, bg=COLOR_TARJETA)
label_ultimo_valor.pack(side="left")

# Contadores OK / NOK
frame_contadores = tk.Frame(card_calidad, bg=COLOR_TARJETA)
frame_contadores.pack(fill="x", pady=8)

label_ok = tk.Label(frame_contadores, text="OK: 0", font=("Helvetica", 16, "bold"),
                    fg="white", bg=COLOR_OK, pady=8)
label_ok.pack(side="left", padx=(0, 8), expand=True, fill="x")

label_nok = tk.Label(frame_contadores, text="NOK: 0", font=("Helvetica", 16, "bold"),
                     fg="white", bg=COLOR_NOK, pady=8)
label_nok.pack(side="left", expand=True, fill="x")

# Reset
btn_reset = tk.Button(card_calidad, text="↺  Reset Contadores", font=("Helvetica", 12),
                      fg=COLOR_TEXTO_SEC, bg="#F1F3F5", bd=0, pady=6,
                      activebackground=COLOR_BORDE, command=reset_contadores)
btn_reset.pack(fill="x", pady=(5, 0))

# ─── Tarjeta 4: Indicadores y Controles ───────────────────────────
card_controles = crear_tarjeta(frame_izquierdo)
card_controles.pack(fill="x", pady=(0, 5))

tk.Label(card_controles, text="CONTROLES", font=("Helvetica", 13, "bold"),
         fg=COLOR_VERDE_WKK, bg=COLOR_TARJETA).pack(pady=(0, 5))

# Indicadores LED + Seguridad (Canvas)
frame_indicadores = tk.Frame(card_controles, bg=COLOR_TARJETA)
frame_indicadores.pack(fill="x", pady=4)

canvas = tk.Canvas(frame_indicadores, width=480, height=60, bg=COLOR_TARJETA,
                   highlightthickness=0)
canvas.pack()

# Indicador Pulso
canvas.create_text(60, 10, text="Pulso", font=("Helvetica", 10, "bold"), fill=COLOR_TEXTO_SEC)
indicador_pulso = canvas.create_oval(28, 20, 92, 55, fill=COLOR_BORDE, outline=COLOR_TEXTO_SEC, width=2)

# Indicador LED
canvas.create_text(180, 10, text="LED", font=("Helvetica", 10, "bold"), fill=COLOR_TEXTO_SEC)
indicador_led = canvas.create_oval(148, 20, 212, 55, fill=COLOR_NOK, outline=COLOR_TEXTO_SEC, width=2)

# Indicador Barrera
canvas.create_text(300, 10, text="Barrera", font=("Helvetica", 10, "bold"), fill=COLOR_TEXTO_SEC)
indicador_barrera = canvas.create_oval(268, 20, 332, 55, fill=COLOR_BORDE, outline=COLOR_TEXTO_SEC, width=2)

# Indicador Paro de Emergencia
canvas.create_text(420, 10, text="Paro Emerg.", font=("Helvetica", 10, "bold"), fill=COLOR_TEXTO_SEC)
indicador_paro = canvas.create_oval(388, 20, 452, 55, fill=COLOR_OK, outline=COLOR_TEXTO_SEC, width=2)

# Etiquetas de estado para Barrera y Paro
frame_estados_seg = tk.Frame(card_controles, bg=COLOR_TARJETA)
frame_estados_seg.pack(fill="x", pady=(0, 4))

tk.Label(frame_estados_seg, text="Barrera:", font=("Helvetica", 11),
         fg=COLOR_TEXTO_SEC, bg=COLOR_TARJETA).pack(side="left", padx=(60, 4))
lbl_barrera_estado = tk.Label(frame_estados_seg, text="---", font=("Helvetica", 11, "bold"),
                              fg=COLOR_TEXTO_SEC, bg=COLOR_TARJETA)
lbl_barrera_estado.pack(side="left", padx=(0, 30))

tk.Label(frame_estados_seg, text="Paro:", font=("Helvetica", 11),
         fg=COLOR_TEXTO_SEC, bg=COLOR_TARJETA).pack(side="left", padx=(0, 4))
lbl_paro_estado = tk.Label(frame_estados_seg, text="---", font=("Helvetica", 11, "bold"),
                           fg=COLOR_TEXTO_SEC, bg=COLOR_TARJETA)
lbl_paro_estado.pack(side="left")

# Botones Encender / Apagar
frame_botones = tk.Frame(card_controles, bg=COLOR_TARJETA)
frame_botones.pack(fill="x", pady=6)

btn_encender = tk.Button(frame_botones, text="● Encender", font=("Helvetica", 13, "bold"),
                         fg="white", bg=COLOR_OK, bd=0, pady=10,
                         activebackground="#00A844", command=encender_led_manual)
btn_encender.pack(side="left", expand=True, fill="x", padx=(0, 6))

btn_apagar = tk.Button(frame_botones, text="■ Apagar", font=("Helvetica", 13, "bold"),
                       fg="white", bg=COLOR_NOK, bd=0, pady=10,
                       activebackground="#C62828", command=apagar_led_manual)
btn_apagar.pack(side="left", expand=True, fill="x", padx=(6, 0))

# Botón Exportar Excel
btn_exportar = tk.Button(card_controles, text="📊  Exportar a Excel", font=("Helvetica", 13, "bold"),
                         fg="white", bg=COLOR_VERDE_WKK, bd=0, pady=10,
                         activebackground=COLOR_VERDE_OSCURO, command=abrir_previsualizacion)
btn_exportar.pack(fill="x", pady=(8, 0))

# ===================================================================
# PANEL DERECHO — Gráfica en tiempo real
# ===================================================================
card_grafica = crear_tarjeta(frame_derecho)
card_grafica.pack(fill="both", expand=True)

tk.Label(card_grafica, text="COMPORTAMIENTO DE FUERZA VS SETPOINTS",
         font=("Helvetica", 13, "bold"), fg=COLOR_VERDE_WKK, bg=COLOR_TARJETA).pack(pady=(0, 6))

figura = Figure(dpi=100)
figura.patch.set_facecolor(COLOR_TARJETA)
ax = figura.add_subplot(111)
ax.set_facecolor("#FAFAFA")
ax.set_xlabel("Muestras en Tiempo Real", fontsize=12, color=COLOR_TEXTO_SEC)
ax.set_ylabel("Fuerza / Presión", fontsize=12, color=COLOR_TEXTO_SEC)
ax.set_ylim(-2, 105) 
ax.grid(True, linestyle="--", alpha=0.3, color="#ADB5BD")

# Bordes limpios y modernos
ax.spines['top'].set_visible(False)
ax.spines['right'].set_visible(False)
ax.spines['left'].set_color(COLOR_BORDE)
ax.spines['bottom'].set_color(COLOR_BORDE)
ax.tick_params(colors=COLOR_TEXTO_SEC, labelsize=10)

linea_fuerza, = ax.plot(datos_fuerza, label="Fuerza Actual", color=COLOR_VERDE_WKK, linewidth=2.5)
linea_umbral, = ax.plot(datos_umbral, label="Umbral Corte", color="#FF9800", linestyle="--", linewidth=1.5)
linea_min, = ax.plot(datos_min, label="Límite Mínimo OK", color="#2196F3", linestyle=":", linewidth=1.5)
linea_max, = ax.plot(datos_max, label="Límite Máximo OK", color="#9C27B0", linestyle=":", linewidth=1.5)
ax.legend(loc="upper left", fontsize=11, framealpha=0.9, edgecolor=COLOR_BORDE)
figura.tight_layout(pad=2)

canvas_grafica = FigureCanvasTkAgg(figura, master=card_grafica)
canvas_grafica.get_tk_widget().pack(fill="both", expand=True)

# --- 8. INICIAR SISTEMA ---
ventana.protocol("WM_DELETE_WINDOW", cerrar)

hilo_modbus = threading.Thread(target=tarea_modbus_alta_velocidad, daemon=True)
hilo_modbus.start()

# Registrar callback del sensor DESPUÉS de que toda la GUI esté lista
time.sleep(0.1)  # Pequeña pausa para estabilizar GPIOs
sensor_pulso.when_pressed = recepcion_pulso
if barrera_conectada:
    barrera.when_pressed = detener_por_seguridad
if paro_conectado:
    paro_emergencia.when_pressed = detener_por_seguridad

refrescar_gui() 
ventana.mainloop()